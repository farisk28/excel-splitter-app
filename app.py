import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
import io
import zipfile
import xml.etree.ElementTree as ET
import msoffcrypto

# =====================================================================
# 1. FUNGSI DEKRIPSI FILE EXCEL TERPROTEKSI PASSWORD
# =====================================================================
def decrypt_excel_file(file_bytes, password=None):
    """
    Mendekripsi file Excel jika diproteksi kata sandi.
    Jika tidak terproteksi, mengembalikan byte data asli.
    """
    file_io = io.BytesIO(file_bytes)
    try:
        office_file = msoffcrypto.OfficeFile(file_io)
        if office_file.is_encrypted():
            if not password:
                return None, "File terkunci kata sandi. Harap masukkan password pada kolom di atas."
            
            decrypted_io = io.BytesIO()
            office_file.load_key(password=password)
            office_file.decrypt(decrypted_io)
            decrypted_io.seek(0)
            return decrypted_io.getvalue(), None
        else:
            return file_bytes, None
    except Exception as e:
        return None, f"Password salah atau gagal membuka file terenkripsi ({str(e)})."

# =====================================================================
# 2. FUNGSI FORMAT NOMINAL & PEMBERSIH TEKS
# =====================================================================
def format_transaction_amount(val):
    if val is None:
        return ""
    val_str = str(val).strip()
    if val_str == "#VALUE!" or not val_str:
        return ""
    try:
        clean_str = val_str.replace(',', '').replace('.', '')
        if val_str.endswith('.0'):
            val_str = val_str[:-2]
            clean_str = val_str.replace(',', '').replace('.', '')
        num = float(clean_str)
        if 0 < num < 1000:
            num = num * 1000
        return f"{int(num):,}"
    except ValueError:
        return val_str

# =====================================================================
# 3. FUNGSI STYLING TABEL & BORDER EXCEL
# =====================================================================
def apply_excel_styling(ws, headers, num_rows):
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True)
    
    # Format Header
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", horizontal="center")

    # Format Data
    for row_idx in range(2, num_rows + 2):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

# =====================================================================
# 4. FUNGSI EKSTRAKSI GAMBAR (IN-CELL & FLOATING)
# =====================================================================
def extract_sheet_images(file_bytes, sheet_name):
    row_images_map = {}
    
    # Ekstraksi In-Cell Image (RichData)
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as z:
            namelist = z.namelist()
            if 'xl/richData/richValueRel.xml' in namelist and 'xl/richData/_rels/richValueRel.xml.rels' in namelist:
                rels_xml = z.read("xl/richData/_rels/richValueRel.xml.rels")
                root_rels = ET.fromstring(rels_xml)
                rel_id_to_target = {e.attrib['Id']: e.attrib['Target'].replace('../media/', '') for e in root_rels}
                
                rich_rel_xml = z.read("xl/richData/richValueRel.xml")
                root_rich_rel = ET.fromstring(rich_rel_xml)
                index_to_media = [rel_id_to_target[e.attrib['{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id']] for e in root_rich_rel if e.attrib['{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'] in rel_id_to_target]

                rv_xml = z.read("xl/richData/rdrichvalue.xml")
                root_rv = ET.fromstring(rv_xml)
                rv_index_to_media = [index_to_media[int(rv[0].text)] for rv in root_rv if int(rv[0].text) < len(index_to_media)]

                sheet_xml = z.read("xl/worksheets/sheet1.xml")
                root_sheet = ET.fromstring(sheet_xml)
                ns_main = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
                sheet_data = root_sheet.find(f'{ns_main}sheetData')
                
                if sheet_data is not None:
                    for row_elem in sheet_data.findall(f'{ns_main}row'):
                        r_idx = int(row_elem.attrib['r'])
                        for c_elem in row_elem.findall(f'{ns_main}c'):
                            vm_attr = c_elem.attrib.get('vm')
                            if vm_attr is not None:
                                vm_idx = int(vm_attr) - 1
                                if 0 <= vm_idx < len(rv_index_to_media):
                                    media_path = f"xl/media/{rv_index_to_media[vm_idx]}"
                                    if media_path in namelist:
                                        img_data = z.read(media_path)
                                        col_idx = 9  # Kolom Evidence default
                                        if r_idx not in row_images_map:
                                            row_images_map[r_idx] = []
                                        row_images_map[r_idx].append((col_idx, io.BytesIO(img_data)))
    except Exception:
        pass

    # Ekstraksi Gambar Floating Biasa
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if hasattr(ws, '_images'):
                for img in ws._images:
                    img_row = img.anchor._from.row + 1
                    img_col = img.anchor._from.col + 1
                    img_bytes = io.BytesIO(img._data())
                    if img_row not in row_images_map:
                        row_images_map[img_row] = []
                    row_images_map[img_row].append((img_col, img_bytes))
    except Exception:
        pass
        
    return row_images_map

# =====================================================================
# 5. FUNGSI MEMBACA MULTI-SHEET EXCEL (DENGAN DUKUNGAN PASSWORD)
# =====================================================================
def process_multisheet_excel(uploaded_files, password=""):
    sheets_dict = {}
    errors = []

    for file in uploaded_files:
        raw_bytes = file.getvalue()
        
        # Jalankan fungsi dekripsi
        decrypted_bytes, err = decrypt_excel_file(raw_bytes, password)
        if err:
            errors.append(f"File **{file.name}**: {err}")
            continue

        try:
            wb = openpyxl.load_workbook(io.BytesIO(decrypted_bytes), data_only=True)
        except Exception as e:
            errors.append(f"Gagal membaca struktur Excel {file.name}: {e}")
            continue

        for s_name in wb.sheetnames:
            ws = wb[s_name]
            current_headers = [str(cell.value).strip() for cell in ws[1] if cell.value is not None]
            
            if not current_headers:
                continue

            if s_name not in sheets_dict:
                sheets_dict[s_name] = {
                    'headers': current_headers,
                    'rows': []
                }

            headers = sheets_dict[s_name]['headers']
            row_images_map = extract_sheet_images(decrypted_bytes, s_name)

            for row_idx in range(2, ws.max_row + 1):
                row_vals = []
                row_dict = {}
                for c in range(1, len(headers) + 1):
                    col_name = headers[c - 1]
                    raw_val = ws.cell(row=row_idx, column=c).value
                    
                    if col_name in ['transaction_amount', 'amount']:
                        val = format_transaction_amount(raw_val)
                    else:
                        val = "" if raw_val is None or str(raw_val).strip() == "#VALUE!" else str(raw_val).strip()
                    
                    row_vals.append(val)
                    row_dict[col_name] = val
                
                row_images = row_images_map.get(row_idx, [])
                if any(v != "" for v in row_vals) or len(row_images) > 0:
                    sheets_dict[s_name]['rows'].append({
                        'data': row_dict,
                        'images': row_images
                    })

    return sheets_dict, errors

# =====================================================================
# 6. FUNGSI MEMBUAT WORKBOOK OUTPUT
# =====================================================================
def create_multisheet_workbook(sheets_dict, target_val, selected_column):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for s_name, s_content in sheets_dict.items():
        headers = s_content['headers']
        rows = s_content['rows']

        if selected_column in headers:
            filtered_rows = [r for r in rows if r['data'].get(selected_column, "") == target_val]
        else:
            filtered_rows = rows

        clean_sheet_name = str(s_name)[:30].replace("/", "_").replace("\\", "_").replace("?", "_")
        ws = wb.create_sheet(title=clean_sheet_name)
        ws.views.sheetView[0].showGridLines = True

        # Header
        for col_idx, h_text in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h_text)

        # Data & Gambar
        for row_offset, item in enumerate(filtered_rows, start=2):
            row_dict = item['data']
            for col_idx, h_text in enumerate(headers, start=1):
                val = row_dict.get(h_text, "")
                ws.cell(row=row_offset, column=col_idx, value=val)

            if item['images']:
                for col_idx, img_bytes in item['images']:
                    img_bytes.seek(0)
                    new_img = Image(img_bytes)
                    new_img.width = 100
                    new_img.height = 100
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws.add_image(new_img, f"{col_letter}{row_offset}")
                    ws.row_dimensions[row_offset].height = 80

        apply_excel_styling(ws, headers, len(filtered_rows))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# =====================================================================
# 7. INTERFACE STREAMLIT
# =====================================================================
st.set_page_config(page_title="Multi-Sheet Excel Splitter", layout="wide")
st.title("📊 Aplikasi Pemisah & Penggabung Berkas Excel")
st.write("Unggah file Excel (termasuk file yang diproteksi password), pilih kolom filter, dan unduh hasilnya secara otomatis!")

col1, col2 = st.columns([2, 1])
with col1:
    uploaded_files = st.file_uploader(
        "Pilih satu atau beberapa file Excel (.xlsx)", 
        type=["xlsx"], 
        accept_multiple_files=True
    )
with col2:
    # FITUR BARU: Input Password
    file_password = st.text_input(
        "🔑 Password File (jika file terenkripsi):", 
        type="password",
        help="Masukkan password jika file Excel yang diunggah terkunci kata sandi."
    )

if uploaded_files:
    with st.spinner("Membaca dan memproses file Excel..."):
        sheets_dict, errors = process_multisheet_excel(uploaded_files, password=file_password)
    
    # Tampilkan pesan error jika password salah atau belum diisi
    if errors:
        for err in errors:
            st.error(err)

    if sheets_dict:
        st.success(f"Berhasil membaca **{len(sheets_dict)}** sheet: {', '.join([f'**{name}** ({len(info['rows'])} baris)' for name, info in sheets_dict.items()])}")

        # Pratinjau Sheet
        st.markdown("### 👀 Pratinjau Data Tiap Sheet")
        tabs = st.tabs(list(sheets_dict.keys()))
        for idx, (s_name, s_info) in enumerate(sheets_dict.items()):
            with tabs[idx]:
                df_preview = pd.DataFrame([r['data'] for r in s_info['rows']])
                st.dataframe(df_preview.head(10))

        st.markdown("---")

        # Kolom Filter
        all_header_sets = [set(info['headers']) for info in sheets_dict.values()]
        common_columns = list(set.intersection(*all_header_sets)) if all_header_sets else []
        if not common_columns:
            common_columns = list(set.union(*all_header_sets))

        st.subheader("⚙️ Atur Filter Pemisahan Data")
        selected_column = st.selectbox(
            "Pilih kolom dasar pemisahan:",
            options=common_columns
        )

        if selected_column:
            all_unique_values = set()
            for s_info in sheets_dict.values():
                if selected_column in s_info['headers']:
                    for r in s_info['rows']:
                        v = r['data'].get(selected_column, "").strip()
                        if v:
                            all_unique_values.add(v)
            
            unique_list = sorted(list(all_unique_values))
            st.info(f"Ditemukan **{len(unique_list)}** kategori unik pada kolom **'{selected_column}'**.")

            st.markdown("---")

            st.subheader("📥 Unduh Hasil Filter")
            if st.button("🚀 Proses & Buat File ZIP Multi-Sheet"):
                with st.spinner("Menyusun file Excel..."):
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        for val in unique_list:
                            excel_buf = create_multisheet_workbook(sheets_dict, val, selected_column)
                            clean_filename = str(val).replace("/", "_").replace("\\", "_").replace("?", "_")
                            zip_file.writestr(f"{clean_filename}.xlsx", excel_buf.getvalue())

                    zip_buffer.seek(0)
                    st.download_button(
                        label=f"📥 Download File ZIP ({len(unique_list)} File Excel)",
                        data=zip_buffer,
                        file_name=f"Hasil_Filter_{selected_column}.zip",
                        mime="application/zip"
                    )
